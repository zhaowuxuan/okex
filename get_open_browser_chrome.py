# import datetime
import json
# from peewee import *
# python -e mysql -u mall_tiny -P -H localhost -p 3306 -s mall_ums > models.py
import logging
import re
import time
import urllib.parse
from decimal import Decimal

import openpyxl
import pypinyin
import requests as s
from fake_useragent import UserAgent
from pyquery import PyQuery as pq
from selenium import webdriver
# from selenium.common.exceptions import TimeoutException
# from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from sqlalchemy import create_engine
from sqlalchemy import text

# from googletrans import Translator

# from selenium.webdriver.support import expected_conditions as EC

chrome_options = webdriver.ChromeOptions()
# 监听
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
# 开发者模式
# chrome_options.add_experimental_option('excludeSwitches', ['enable-automation'])
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
# 无窗口模式
# chrome_options.add_argument('--headless')
# chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile"

# 全局配置
logging.basicConfig(level=logging.DEBUG)
URL = "http://www.iwencai.com/unifiedwap/result?w={0}&querytype=&issugs"
QUERY = "最新股东人数；前十大流通股占比；流通市值；总市值；总股本；流通股本；股息；每股净收益；净收益；股东权益；净利润；销售收入；资产总额；净资产"
QUERY = urllib.parse.quote(QUERY)
URL = URL.format(QUERY)
# 共享数据
# 汇总数据
G_DATA = list()
# 全局标题
G_TITLES = list()
# 全局en标题
G_EN_TITLES = list()
# 全局py标题
G_PY_TITLES = list()
# 全局页码列表
G_PAGES = list()
# 全局行号
G_ROW = 1
# 全局頁碼
G_CUR_PAGE = 1
# 当页数据
G_CUR_PAGE_DATA = list()
# 当前行数据
G_CUR_ROW_DATA = list()
# 代码
CODE = 'dm'
# 代码列索引
DM_IDX = 0
# 驱动初始化
ENGINE = create_engine('mysql+pymysql://mall_tiny:%s@localhost:3306/mall_ums' % urllib.parse.quote('1233rs!@d'),
                       echo=False)
# 全局定义
MAIN_TABLE = "share_information"
DATE_TABLE = "share_date_information"
DATA_DATE_COL = 'data_date'
COL_DEF_TABLE = {'TABLE': 'share_data_col_def', 'COL_DEF': 'col_def', 'COL_DESC': 'col_desc'}
REG_BLANK_SPACE = "r'\s+'"
REG_BLANK_BRACKET = "u'\\(.*?\\)|\\{.*?}|\\[.*?]'"


def format_date(p_date):
    """
    判断是否是一个有效的日期字符串
    :param p_date:
    :return:
    """
    try:
        return time.strftime("%Y%m%d", time.strptime(p_date, "%Y%m%d"))
    except ValueError:
        try:
            return time.strftime("%Y%m%d", time.strptime(p_date, "%Y-%m-%d"))
        except ValueError:
            try:
                return time.strftime("%Y%m%d", time.strptime(p_date, "%Y/%m/%d"))
            except ValueError:
                try:
                    return time.strftime("%Y%m%d", time.strptime(p_date, "%Y.%m.%d"))
                except ValueError:
                    return False


def valid_date(p_date):
    """
    判断是否是一个有效的日期字符串
    :param p_date:
    :return:
    """
    try:
        time.strptime(p_date, "%Y%m%d")
        # p_type = 2
        return 1
    except ValueError:
        try:
            # p_type = 1
            time.strptime(p_date, "%Y-%m-%d")
            return 2
        except ValueError:
            try:
                # p_type = 3
                time.strptime(p_date, "%Y/%m/%d")
                return 3
            except ValueError:
                try:
                    # p_type = 4
                    time.strptime(p_date, "%Y.%m.%d")
                    return 4
                except ValueError:
                    return False


# 全局doc
# doc = ''
def get_pinyin(p_titles):
    """
    获取中文拼音首字母
    :param p_titles:
    :return:
    """
    global DM_IDX
    for title in p_titles:
        # 删除括号内容
        py = "".join(
            pypinyin.lazy_pinyin((re.sub(u"\\(.*?\\)|\\{.*?}|\\[.*?]", "", title)), style=pypinyin.Style.FIRST_LETTER))
        G_PY_TITLES.append(py)
    # 获取代码列位置
    DM_IDX = G_PY_TITLES.index(CODE)
    logging.info(G_PY_TITLES)


def update_table(p_table_name, p_id, p_cols, p_cols_data, p_date):
    """
    更新指定表指定列数据
    :param p_date:
    :param p_cols_data:
    :param p_table_name:
    :param p_id:
    :param p_cols:
    :return:
    """
    # global ENGINE
    # dm_idx = p_cols.index(CODE)
    data = list(map(turn_to_mysql_data, p_cols_data, p_cols))
    logging.debug('update_table cols= ' + '(' + ','.join(p_cols) + ')')
    # 代码列数据会被误识别为数字，暂做覆盖处理 BUG(更新时未使用p_col_data中的dm)
    # data[dm_idx] = "'" + p_cols_data[dm_idx] + "'"
    logging.debug('update_table values= ' + '(' + ','.join(data) + ')')
    set_str = ','.join(list(map(lambda x, y: str(x) + " = " + str(y), p_cols, data)))
    logging.debug('update_table set_str=' + set_str)
    if p_date:
        sql = "update {0} set {1} where dm = '{2}' and data_date = str_to_date('{3}','%Y%m%d');".format(
            p_table_name, set_str, p_id, p_date)
    else:
        sql = "update {0} set {1} where dm = '{2}';".format(p_table_name, set_str,
                                                            p_id)
    with ENGINE.connect() as connection:
        result = connection.execute(text(sql))
        logging.debug("update_table sql=" + sql)
        return result


def insert_table(p_table_name, p_cols, p_cols_data):
    """
    插入指定表指定列
    :param p_table_name:
    :param p_cols:
    :param p_cols_data:
    :return:
    """
    dm_idx = p_cols.index(CODE)
    # global ENGINE
    data = list(map(turn_to_mysql_data, p_cols_data, p_cols))
    logging.debug('insert_table cols= ' + '(' + ','.join(p_cols) + ')')
    # 覆盖原代码值
    data[dm_idx] = "'" + p_cols_data[dm_idx] + "'"
    logging.debug('insert_table values= ' + '(' + ','.join(data) + ')')
    sql = "insert into {0} {1} values {2};".format(p_table_name, '(' + ','.join(p_cols) + ')',
                                                   '(' + ','.join(data) + ')')
    with ENGINE.connect() as connection:
        result = connection.execute(text(sql))
        logging.debug("insert_table sql=" + sql)
    return result


def add_table_col(p_table_name, p_col_name, p_type, p_comment):
    """
    增加指定表指定列
    :param p_table_name:
    :param p_col_name:
    :param p_type:
    :param p_comment:
    :return:
    """
    # global ENGINE
    sql = "alter table {0} add column {1} {2} comment '{3}';".format(p_table_name, p_col_name, p_type, p_comment)
    logging.debug('add_table_col sql =' + sql)

    with ENGINE.connect() as connection:
        result = connection.execute(text(sql))
    return result


def add_col_desc(p_table_dict: dict, p_col_name: str, p_comment: str):
    """
    记录列描述
    :param p_table_dict:
    :param p_col_name:
    :param p_comment:
    :return:
    """
    with ENGINE.connect() as connection:
        sql = "insert into {0} ({1},{2}) values ('{3}','{4}');".format(p_table_dict['TABLE'],
                                                                       p_table_dict['COL_DEF'],
                                                                       p_table_dict['COL_DESC'], p_col_name,
                                                                       p_comment)
        result = connection.execute(text(sql))
        logging.debug("add_col_desc sql=" + sql)
    return result


def turn_to_mysql_data(p_col_data: str, p_col: str):
    """
    返回字符串对应的MYSQL数据库可识别类型
    :param p_col:
    :param p_col_data:
    :return:
    """
    data_type = get_mysql_data_type(p_col_data)
    logging.info(data_type)
    if 'int' in data_type:
        return amt_format(p_col_data, 'V')
    elif 'date' == data_type:
        if valid_date(p_col_data) == 2:
            return "str_to_date('{0}','%Y-%m-%d')".format(p_col_data)
        elif valid_date(p_col_data) == 3:
            return "str_to_date('{0}','%Y/%m/%d')".format(p_col_data)
        elif valid_date(p_col_data) == 4:
            return "str_to_date('{0}','%Y.%m.%d')".format(p_col_data)
        else:
            return "str_to_date('{0}','%Y%md%d')".format(p_col_data)
    elif 'decimal' in data_type:
        return p_col_data
    # 如果内容为--，则返回其他记录对应类型
    elif '-' in p_col_data and p_col_data.strip('-') == '':
        for i in range(len(G_DATA)):
            return turn_to_mysql_data(G_DATA[i][G_PY_TITLES.index(p_col)], p_col)
    else:
        return "'" + str(p_col_data) + "'"


def get_mysql_data_type(p_col_data: str):
    """
    获取对应数据的MYSQL类型
    :param p_col_data:
    :return:
    """
    # BUG 20210101 等类日期数值无法判断
    if p_col_data.isdecimal():
        if len(p_col_data) > 9:
            return "bigint"
        else:
            return "int"
    elif valid_date(p_col_data):
        return "date"
    else:
        try:
            float(p_col_data)
            return "decimal({0},{1})".format(10, 2)
        except ValueError:
            try:
                return amt_format(p_col_data, 'T')
            except ValueError:
                return "varchar({0})".format(len(p_col_data) * 5)


def amt_format(p_amt: str, p_type):
    """
    格式化金额
    :param p_amt:
    :param p_type:
    :return:
    """
    if p_type == 'T':
        if p_amt[-1:] == '亿' and float(p_amt[0:-1].replace(',', '')):
            # 数字
            if len(str(Decimal(re.sub(r'[,]', '', p_amt[0:-1])) * 100000000)) > 9:
                return "bigint"
            else:
                return "int"
        elif p_amt[-1:] == '万' and float(p_amt[0:-1].replace(',', '')):
            # 数字
            if len(str(Decimal(re.sub(r'[,]', '', p_amt[0:-1])) * 10000)) > 9:
                return "bigint"
            else:
                return "int"
        else:
            if len(str(float(re.sub(r'[,]', '', p_amt)))) > 9:
                return "bigint"
            else:
                return "int"
    elif p_type == 'V':
        if p_amt[-1:] == '亿' and float(p_amt[0:-1].replace(',', '')):
            return str(Decimal(re.sub(r'[,]', '', p_amt[0:-1])) * 100000000)
        elif p_amt[-1:] == '万' and float(p_amt[0:-1].replace(',', '')):
            return str(Decimal(re.sub(r'[,]', '', p_amt[0:-1])) * 10000)
        else:
            return str(Decimal(re.sub(r'[,]', '', p_amt)))


def is_col_exists(p_table_name, p_col_name):
    """
    判断表内列是否存在
    :param p_table_name:
    :param p_col_name:
    :return:
    """
    with ENGINE.connect() as connection:
        result = connection.execute(text("select table_name from information_schema.columns where \
        table_schema = 'mall_ums' and table_name = '{0}'  and column_name='{1}';".format(p_table_name, p_col_name)))
        # for row in result:
        #     logging.info("username:", row['username'])
        return result.rowcount


def is_id_exists(p_table_name, p_dm):
    """
    判断表内某ID是否存在
    :param p_table_name:
    :param p_dm:
    :return:
    """
    with ENGINE.connect() as connection:
        result = connection.execute(text("select dm from {0} where \
        dm = '{1}' limit 1;".format(p_table_name, p_dm)))
        # for row in result:
        #     logging.info("username:", row['username'])
        return result.rowcount


def save_data_to_mysql(p_titles, p_py_titles: list, p_data: list):
    """
    将数据载入MYSQL
    :param p_py_titles:
    :param p_titles:
    :param p_data:
    :return:
    """
    main_index = list()
    date_index = dict()
    global G_CUR_ROW_DATA
    # 遍历标题
    for title in p_py_titles:
        # 取标题位置
        title_index = p_py_titles.index(title)
        # 标题中有空格
        if ' ' in title:
            # 分割标题
            del_title = str(title).split(' ')
            del_zh_title = p_titles[title_index].split(' ')
            logging.info(del_title)
            logging.info(del_title[len(del_title) - 1])
            # 标题中有日期
            datestr = format_date(del_title[len(del_title) - 1])
            if datestr:
                # 记录有日期的索引
                # date_index append({valid_date(del_title[len(del_title) - 1]), title_index})
                if date_index.get(datestr):
                    date_index[datestr].append(title_index)
                else:
                    date_index.setdefault(datestr, [title_index])
                # 修改标题
                p_py_titles[title_index] = del_title[0]
                p_titles[title_index] = del_zh_title[0]
            else:
                # 只修改标题
                p_py_titles[title_index] = del_title[0]
                p_titles[title_index] = del_zh_title[0]

            if not is_col_exists(DATE_TABLE, p_py_titles[title_index]):
                # 修改表结构
                if p_data[0][title_index]:
                    # 获取数据库类型
                    dt = get_mysql_data_type(p_data[0][title_index])
                    logging.info(dt)
                    # 新增表列
                    add_table_col(DATE_TABLE, p_py_titles[title_index], dt, p_titles[title_index])
                    # 新增列描述
                    add_col_desc(COL_DEF_TABLE, p_py_titles[title_index], p_titles[title_index])
                else:
                    logging.info('跳过：' + p_py_titles[title_index])
                    pass
        else:
            # 记录无日期的索引
            main_index.append(title_index)
            # 列不存在
            if not is_col_exists(MAIN_TABLE, p_py_titles[title_index]):
                # 修改表结构
                if p_data[0][title_index]:
                    dt = get_mysql_data_type(p_data[0][title_index])
                    logging.info(dt)
                    # 新增表列
                    add_table_col(MAIN_TABLE, p_py_titles[title_index], dt, p_titles[title_index])
                    # 新增列描述
                    add_col_desc(COL_DEF_TABLE, p_py_titles[title_index], p_titles[title_index])
                else:
                    logging.info('跳过：' + p_py_titles[title_index])
                    pass

    logging.info(date_index)
    logging.info("title end")
    dm_index = p_py_titles.index(CODE)
    # first_flag = True

    # 按列
    # main_data_col.append(p_data['dm'])
    # date_data_col.append(p_data['dm'])
    # for i in main_index:
    #     main_data_col.append(p_data[i])
    # for j in date_index:
    #     date_data_col.append(p_data[j])
    # 分离日期数据和固定数据并更新
    # 取一行数据
    for data in p_data:
        # 初始化
        G_CUR_ROW_DATA = data
        main_data_row = list()
        main_data_title_row = list()
        # 数据分类
        for i in main_index:
            # 空数据跳过
            if data[i]:
                main_data_row.append(data[i])
                main_data_title_row.append(p_py_titles[i])
            else:
                logging.info('跳过：' + p_py_titles[i])
                pass

        # 主表
        if is_id_exists(MAIN_TABLE, data[dm_index]):
            # 更新
            pass
            # update_table(MAIN_TABLE, data[dm_index], main_data_title_row, main_data_row, None)
        else:
            insert_table(MAIN_TABLE, main_data_title_row, main_data_row)

        # 日期表
        if is_id_exists(DATE_TABLE, data[dm_index]):
            for j in date_index:
                # 日期表
                date_data_row = list()
                date_data_title_row = list()
                for k in date_index[j]:
                    if data[k]:
                        date_data_row.append(data[k])
                        date_data_title_row.append(p_py_titles[k])
                        logging.info("title=" + p_py_titles[k] + "    date=" + data[k])
                    else:
                        logging.info('跳过：' + p_py_titles[k])
                        pass
                # 更新
                update_table(DATE_TABLE, data[dm_index], date_data_title_row, date_data_row, j)
        else:
            for j in date_index:
                # 日期表
                # 初始化日期和代码
                date_data_row = list()
                date_data_title_row = list()
                date_data_row.append(data[dm_index])
                date_data_title_row.append(p_py_titles[dm_index])
                # 写死日期列
                date_data_row.append(j)
                date_data_title_row.append(DATA_DATE_COL)
                for k in date_index[j]:
                    if data[k]:
                        date_data_row.append(data[k])
                        date_data_title_row.append(p_py_titles[k])
                        logging.info("title=" + p_py_titles[k] + "    date=" + data[k])
                    else:
                        logging.info('跳过：' + p_py_titles[k])
                        pass
                # 插入
                insert_table(DATE_TABLE, date_data_title_row, date_data_row)


def translate(p_text):
    """
    翻译API调用
    :param p_text:
    :return:
    """
    headers = {'X-Forwarded-For': '', 'User-Agent': (UserAgent()).random}
    data = {"f": "auto", "t": "auto", "w": p_text}
    icb_url = "http://fy.iciba.com/ajax.php?a=fy"
    response = s.post(icb_url, headers=headers, data=data, timeout=15).text
    logging.info(response)
    obj = json.loads(response)
    time.sleep(1)
    status = obj.get('status')
    content = obj.get('content')
    if status == 1:
        out = content.get('out')
        return out
    else:
        word_mean, out = "".join(content.get("word_mean")), ''
        mean = re.findall(r";(.*?);", word_mean)
        if not mean:
            mean = re.findall(r"(.*?);", word_mean)
        if mean:
            out = mean[0]
    return out


def get_en_title(p_titles):
    """
    获取英文
    :param p_titles:
    :return:
    """
    for title in p_titles:
        G_EN_TITLES.append(translate(title))
    logging.info(G_EN_TITLES)


def get_title_data(p_doc, p_wb):
    """
    获取web标题数据
    :param p_doc:
    :param p_wb:
    :return:
    """
    global G_TITLES
    # 取标题
    # title = p_doc('.app-container .main-body.set-height .result-page-wrapper.mainBox .main-content-x .table-wrapper .iwc-table-wrapper-outer .iwc-table-wrapper.clearfix .iwc-table-container .iwc-table-content.isTwoLine .iwc-table-scroll .iwc-table-header.table-right-thead.scroll-style2 .iwc-table-header-ul.clearfix').find('span')
    t_titles = p_doc(
        '.app-container .main-body.set-height .result-page-wrapper.mainBox .main-content-x .table-wrapper .iwc-table-wrapper-outer .iwc-table-wrapper.clearfix .iwc-table-container .iwc-table-content.isTwoLine .iwc-table-scroll .iwc-table-header.table-right-thead.scroll-style2 .iwc-table-header-ul.clearfix').children()
    t_sub_titles = ["序号", "选择", "代码", "名称"]
    # logging.info(type(titles))
    # logging.info(type(titles.items()))
    G_TITLES = t_sub_titles
    for title in t_titles.items():
        G_TITLES.append(title.find('span').text())
    # 打印标题
    logging.info(G_TITLES)
    sheet = p_wb.active
    k = 0
    for title in G_TITLES:
        c1 = sheet.cell(row=1, column=k + 1)
        c1.value = str(title)
        k += 1


def get_data(p_doc, p_wb):
    """
    获取web表内数据
    :param p_doc:
    :param p_wb:
    :return:
    """
    page_data = list()
    data_to_str = list()
    global G_ROW
    global G_DATA
    global G_CUR_PAGE_DATA
    G_CUR_PAGE_DATA = list()
    # 取数据
    doc_page_data = p_doc(
        '.app-container .main-body.set-height .result-page-wrapper.mainBox .main-content-x .table-wrapper .iwc-table-wrapper-outer .iwc-table-wrapper.clearfix .iwc-table-container .iwc-table-content.isTwoLine .iwc-table-scroll .iwc-table-body.scroll-style2.big-mark table tr')
    for doc_row_data in doc_page_data.items():
        row_data = list()
        for col_data in doc_row_data.find('td').items():
            row_data.append(col_data.text())
        page_data.append(row_data)
        data_to_str = list(map(str, row_data))
        G_DATA.append(data_to_str)
        G_CUR_PAGE_DATA.append(data_to_str)

    save_data_to_mysql(G_TITLES, G_PY_TITLES, G_CUR_PAGE_DATA)
    # logging.info(row_data)
    # logging.info(page_data)

    sheet = p_wb.active
    for row_data in page_data:
        k = 0
        G_ROW += 1
        # logging.info('G_ROW' + str(G_ROW))
        for col_data in row_data:
            c1 = sheet.cell(row=G_ROW, column=k + 1)
            c1.value = str(col_data)
            k += 1


def max_page_number(p_doc):
    """
    获取web最大页码
    :param p_doc:
    :return:
    """
    global G_PAGES
    t_pages = p_doc('.pcwencai-pagination li').children().items()
    # logging.info(t_pages)
    for p in t_pages:
        # logging.info(p.text())
        G_PAGES.append(p.text())
    G_PAGES.remove('上页')
    G_PAGES.remove('下页')
    G_PAGES.remove('…')
    # logging.info("max_page_number 最大页码=" + str(max(G_PAGES)))
    return max(G_PAGES)


def main():
    """
    入口程序
    :return:
    """
    global G_CUR_PAGE
    # 启动驱动程序
    with webdriver.Chrome(options=chrome_options,
                          executable_path="G:\WorkSpace\Python\MyTest\snow\chromedriver.exe") as browser:
        browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                        Object.defineProperty(navigator, 'webdriver', {
                          get: () => undefined
                        })
                      """
        })
        wait = WebDriverWait(browser, 30)
        # 创建excel操作对象
        wb = openpyxl.Workbook()
        # 打开网址
        browser.get(URL)
        time.sleep(5)
        # 获取网页代码
        # html = browser.page_source
        # logging.info(html)
        # 获取网页pq对象
        doc = pq(browser.page_source)
        # browser.find_elements_by_css_selector("page-item")
        # logging.info(doc('.pcwencai-pagination li').children().text())
        # logging.info(browser.find_element_by_css_selector(".page-item.active").text)
        # 等待测试代码
        # WebDriverWait(browser, 10, 0.5, ignored_exceptions=TimeoutException).until(lambda x: x.find_element_by_css_selector(".page-item.active").text == str(1))

        # 获取标题数据
        get_title_data(doc, wb)
        get_pinyin(G_TITLES)
        # 获取当前页数据
        get_data(doc, wb)
        # 最大页码
        mp = int(max_page_number(doc))
        logging.info("mp 最大页码=" + str(mp))

        # 翻页
        while mp > 1 and int(G_CUR_PAGE) < mp:
            # 换页
            # 此方法无法触发按钮，下页元素被覆盖
            # browser.find_element_by_link_text('下页').click()
            next_page = browser.find_element_by_link_text('下页')

            # 通过js触发事件
            browser.execute_script("arguments[0].click();", next_page)
            # logging.info('点击下一页后，第' + doc('.page-item.active').children().text() + '页')

            # 等待table加载完成
            # logging.info(browser.find_element_by_css_selector(".page-item.active").text)
            WebDriverWait(browser, 10).until(
                lambda x: x.find_element_by_css_selector(".page-item.active").text != G_CUR_PAGE)
            logging.info('等待加载后，第' + browser.find_element_by_css_selector(".page-item.active").text + '页')
            # logging.info(datetime.datetime.now())
            G_CUR_PAGE = browser.find_element_by_css_selector(".page-item.active").text
            logging.info('G_CUR_PAGE 第' + G_CUR_PAGE + '页')

            # 不加延时，G_DATA中会重复第一页的数据（原因不明,pq函数异步处理？）
            time.sleep(5)
            # 获取网页pq对象
            doc = pq(browser.page_source)

            # 获取当前页数据
            get_data(doc, wb)
        # 数据持久化
        # save_data_to_mysql(G_TITLES, G_PY_TITLES, G_DATA)
        wb.save("./demo1.xlsx")
        # 设置等待
        wait = WebDriverWait(browser, 10)


if __name__ == '__main__':
    main()
